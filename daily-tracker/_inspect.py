"""Temporary script to inspect Excel file structures."""
import openpyxl
from datetime import datetime, date

files_to_check = [
    ("D:/my college/zyc学习计划/博一上+寒假.xlsx", "日历"),
    ("D:/my college/zyc学习计划/大三上工作计划.xlsx", None),
]

for fpath, sheet_name in files_to_check:
    print(f"\n{'='*60}")
    print(f"FILE: {fpath}")
    wb = openpyxl.load_workbook(fpath, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    target_sheet = sheet_name if sheet_name else wb.sheetnames[0]
    sheet = wb[target_sheet]
    print(f"Sheet '{target_sheet}': {sheet.max_row} rows x {sheet.max_column} cols")

    for r in range(1, min(15, sheet.max_row + 1)):
        parts = []
        for c in range(1, min(sheet.max_column + 1, 11)):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                if isinstance(val, (datetime, date)):
                    parts.append(f"c{c}=[DATE:{val}]")
                else:
                    parts.append(f"c{c}={str(val)[:35]}")
        if parts:
            print(f"  Row {r:2d}: {' | '.join(parts)}")
        else:
            print(f"  Row {r:2d}: (empty)")
    wb.close()

# Check 大四下+暑假.xlsx which has month sheets
print(f"\n{'='*60}")
fpath = "D:/my college/zyc学习计划/大四下+暑假.xlsx"
wb = openpyxl.load_workbook(fpath, data_only=True)
print(f"FILE: {fpath}")
print(f"Sheets: {wb.sheetnames}")
for sn in wb.sheetnames[:2]:
    sheet = wb[sn]
    print(f"\nSheet '{sn}': {sheet.max_row} rows x {sheet.max_column} cols")
    for r in range(1, min(10, sheet.max_row + 1)):
        parts = []
        for c in range(1, min(sheet.max_column + 1, 11)):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                if isinstance(val, (datetime, date)):
                    parts.append(f"c{c}=[DATE:{val}]")
                else:
                    parts.append(f"c{c}={str(val)[:35]}")
        if parts:
            print(f"  Row {r:2d}: {' | '.join(parts)}")
wb.close()
