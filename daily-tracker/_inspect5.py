"""Check formula files with data_only=True."""
import openpyxl
from datetime import datetime, date

fpath = "D:/my college/zyc学习计划/大三上工作计划.xlsx"
wb = openpyxl.load_workbook(fpath, data_only=True)
sheet = wb[wb.sheetnames[0]]
print(f"Sheet: {sheet.title}, rows={sheet.max_row}, cols={sheet.max_column}")

for r in range(1, 25):
    date_count = 0
    parts = []
    for c in range(1, min(sheet.max_column + 1, 11)):
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            if isinstance(val, (datetime, date)):
                parts.append(f"c{c}=[DATE:{val}]")
                date_count += 1
            else:
                s = str(val)[:30]
                parts.append(f"c{c}={s}")
    if parts:
        marker = " *** DATE ROW" if date_count >= 3 else ""
        print(f"  Row {r:2d}: {' | '.join(parts)}{marker}")
wb.close()
