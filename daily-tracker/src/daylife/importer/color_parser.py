"""颜色解析模块 - 从openpyxl单元格提取颜色并映射到任务状态

颜色映射规则：
- 灰色系背景 → status=completed
- 红色系背景或红色字体 → status=incomplete
- 其他/无颜色 → status=in_progress
"""

from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color


def _color_to_rgb(color: Color | None) -> tuple[int, int, int] | None:
    """将openpyxl Color对象转换为(R, G, B)元组。

    openpyxl的颜色可能是：
    - theme颜色（需要特殊处理）
    - indexed颜色（调色板索引）
    - rgb字符串（如 'FFRRGGBB' 或 'RRGGBB'）
    """
    if color is None:
        return None

    # 检查是否为 rgb 格式
    if color.rgb and isinstance(color.rgb, str) and color.rgb != "00000000":
        rgb_str = color.rgb
        # openpyxl rgb 格式为 AARRGGBB (带alpha通道) 或 RRGGBB
        if len(rgb_str) == 8:
            # AARRGGBB - 跳过alpha
            r, g, b = int(rgb_str[2:4], 16), int(rgb_str[4:6], 16), int(rgb_str[6:8], 16)
        elif len(rgb_str) == 6:
            r, g, b = int(rgb_str[0:2], 16), int(rgb_str[2:4], 16), int(rgb_str[4:6], 16)
        else:
            return None
        return r, g, b

    # indexed颜色 - 常见索引映射
    if color.indexed is not None and color.indexed != 64:  # 64 = system default
        # openpyxl 内置的 indexed 颜色表
        _INDEXED_COLORS = {
            0: (0, 0, 0),        # Black
            1: (255, 255, 255),  # White
            2: (255, 0, 0),      # Red
            3: (0, 255, 0),      # Green
            4: (0, 0, 255),      # Blue
            5: (255, 255, 0),    # Yellow
            6: (255, 0, 255),    # Magenta
            7: (0, 255, 255),    # Cyan
            8: (0, 0, 0),        # Black
            9: (255, 255, 255),  # White
            22: (192, 192, 192), # Silver/Gray
            23: (128, 128, 128), # Gray
            55: (128, 128, 128), # Gray
        }
        return _INDEXED_COLORS.get(color.indexed)

    # theme颜色 - 常见 theme 索引
    if color.theme is not None:
        # theme 0=白, 1=黑, 2-9=强调色, 通常无法精确还原
        # 但 tint 可以判断明暗度
        _THEME_BASE = {
            0: (255, 255, 255),  # lt1 (white)
            1: (0, 0, 0),        # dk1 (black)
            2: (68, 84, 106),    # lt2
            3: (231, 230, 230),  # dk2
            4: (68, 114, 196),   # accent1
            5: (237, 125, 49),   # accent2
            6: (165, 165, 165),  # accent3 (gray)
            7: (255, 192, 0),    # accent4
            8: (91, 155, 213),   # accent5
            9: (112, 173, 71),   # accent6
        }
        base = _THEME_BASE.get(color.theme)
        if base:
            # 应用 tint（正值变亮，负值变暗）
            tint = color.tint or 0.0
            if tint > 0:
                r = int(base[0] + (255 - base[0]) * tint)
                g = int(base[1] + (255 - base[1]) * tint)
                b = int(base[2] + (255 - base[2]) * tint)
            elif tint < 0:
                r = int(base[0] * (1 + tint))
                g = int(base[1] * (1 + tint))
                b = int(base[2] * (1 + tint))
            else:
                r, g, b = base
            return (max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b)))

    return None


def is_gray(r: int, g: int, b: int, tolerance: int = 40) -> bool:
    """判断RGB是否为灰色系。

    灰色特征：R≈G≈B，且不是纯白(255,255,255)也不是纯黑(0,0,0)。
    """
    # 排除纯白和接近纯白
    if r > 240 and g > 240 and b > 240:
        return False
    # 排除纯黑和接近纯黑
    if r < 15 and g < 15 and b < 15:
        return False
    # R,G,B 三值接近
    avg = (r + g + b) / 3
    return (abs(r - avg) < tolerance and
            abs(g - avg) < tolerance and
            abs(b - avg) < tolerance)


def is_red(r: int, g: int, b: int) -> bool:
    """判断RGB是否为红色系。

    红色特征：R 显著高于 G 和 B。
    """
    return r > 150 and g < 120 and b < 120 and r > g + 50


def get_cell_status(cell: Cell) -> str:
    """从单元格的背景色和字体颜色推断任务状态。

    Returns:
        'completed' / 'incomplete' / 'in_progress'
    """
    # 1. 检查背景色
    fill: PatternFill = cell.fill
    if fill and fill.patternType and fill.patternType != "none":
        bg_rgb = _color_to_rgb(fill.fgColor)
        if bg_rgb:
            if is_gray(*bg_rgb):
                return "completed"
            if is_red(*bg_rgb):
                return "incomplete"

    # 2. 检查字体颜色
    font: Font = cell.font
    if font and font.color:
        font_rgb = _color_to_rgb(font.color)
        if font_rgb and is_red(*font_rgb):
            return "incomplete"

    # 3. 检查删除线（有些用户用删除线标记完成）
    if font and font.strikethrough:
        return "completed"

    return "in_progress"


def describe_color(cell: Cell) -> str:
    """返回单元格颜色的可读描述（用于调试/预览）。"""
    parts = []

    fill: PatternFill = cell.fill
    if fill and fill.patternType and fill.patternType != "none":
        bg_rgb = _color_to_rgb(fill.fgColor)
        if bg_rgb:
            r, g, b = bg_rgb
            parts.append(f"bg=#{r:02X}{g:02X}{b:02X}")
            if is_gray(r, g, b):
                parts.append("(gray)")
            elif is_red(r, g, b):
                parts.append("(red)")

    font: Font = cell.font
    if font and font.color:
        font_rgb = _color_to_rgb(font.color)
        if font_rgb:
            r, g, b = font_rgb
            parts.append(f"font=#{r:02X}{g:02X}{b:02X}")

    if font and font.strikethrough:
        parts.append("strikethrough")

    return " ".join(parts) if parts else "no-color"
