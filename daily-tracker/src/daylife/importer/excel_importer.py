"""Excel历史数据导入模块

支持批量导入用户7-8年的Excel学习计划记录到DayLife数据库。

用法：
    python -m src.importer.excel_importer --path 'D:/my college/zyc学习计划/' --dry-run
    python -m daylife.importer.excel_importer --path 'D:/my college/zyc学习计划/'
"""

import argparse
import json
import logging
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from daylife.importer.base import BaseImporter, ImportResult
from daylife.importer.color_parser import describe_color, get_cell_status
from daylife.importer.date_corrector import (
    correct_date,
    parse_date_from_cell,
    parse_semester_from_filename,
)

logger = logging.getLogger(__name__)

# 自动分类关键词映射
CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "学习": ["学习", "课程", "上课", "复习", "预习", "作业", "考试", "阅读", "读书", "笔记", "英语", "数学"],
    "科研": ["科研", "论文", "实验", "课题", "研究", "文献", "综述", "数据分析", "仿真", "模型",
             "paper", "experiment", "research"],
    "编程": ["编程", "代码", "开发", "debug", "项目", "python", "java", "code", "git", "程序"],
    "运动": ["运动", "跑步", "晨跑", "健身", "篮球", "游泳", "骑车", "锻炼", "体育", "羽毛球", "乒乓球"],
    "生活": ["生活", "购物", "打扫", "洗衣", "做饭", "吃饭", "理发", "取快递", "超市", "医院"],
    "社交": ["社交", "聚餐", "约会", "见面", "电话", "聊天", "朋友", "同学", "老师"],
    "工作": ["工作", "兼职", "实习", "面试", "简历", "招聘", "工资", "公司"],
    "娱乐": ["娱乐", "游戏", "电影", "电视", "刷剧", "音乐", "旅游", "出游", "逛街"],
    "休息": ["休息", "睡觉", "午休", "放假", "摸鱼"],
}


def _is_pure_number(text: str) -> bool:
    """检查文本是否为纯数字（可能是年份、月份等非内容值）。"""
    try:
        float(text)
        return True
    except ValueError:
        return False


# 应跳过的文本（日历模板、表头、星期标识等）
_SKIP_TEXTS = {
    "备注", "备注：", "备注:", "notes", "note",
    "星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日",
    "mon", "tue", "wed", "thu", "fri", "sat", "sun",
    "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
    "家长教师会议（下午 7 点）",  # Excel日历模板默认内容
}

# 月份标识
_MONTH_PATTERN = re.compile(r"^\d{1,2}月$")


def _is_skip_text(text: str) -> bool:
    """检查文本是否应跳过（模板占位符、表头、星期等）。"""
    lower = text.strip().lower()
    if lower in _SKIP_TEXTS or text.strip() in _SKIP_TEXTS:
        return True
    if _MONTH_PATTERN.match(text.strip()):
        return True
    return False


def auto_categorize(content: str) -> str:
    """根据内容关键词自动分类。"""
    content_lower = content.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in content_lower:
                return category
    return "其他"


def _detect_calendar_grid(sheet: Worksheet) -> dict | None:
    """检测日历网格格式：日期行 + 内容行交替排列，跨7列（一周）。

    扫描整个sheet，找出所有日期行（一行中有>=3个日期值的行）。

    Returns:
        {"date_rows": {row_idx: {col_idx: date, ...}, ...}, "col_range": (start, end)}
        或 None 如果不是日历网格
    """
    date_rows: dict[int, dict[int, date]] = {}

    for row_idx in range(1, sheet.max_row + 1):
        row_dates: dict[int, date] = {}
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None:
                continue
            if isinstance(val, datetime):
                row_dates[col_idx] = val.date()
            elif isinstance(val, date):
                row_dates[col_idx] = val
        # 如果一行有>=3个日期值，认为是日期行
        if len(row_dates) >= 3:
            date_rows[row_idx] = row_dates

    if len(date_rows) >= 2:
        # 确定列范围
        all_cols = set()
        for rd in date_rows.values():
            all_cols.update(rd.keys())
        return {
            "date_rows": date_rows,
            "col_range": (min(all_cols), max(all_cols)),
        }
    return None


def _detect_weekday_grid(sheet: Worksheet) -> dict | None:
    """检测星期网格格式：表头为 day1-day7 或 Mon-Sun，无日期。

    Returns:
        {"header_row": int, "col_range": (start, end), "days_per_row": int}
        或 None
    """
    weekday_patterns = {
        "day1", "day2", "day3", "day4", "day5", "day6", "day7",
        "mon", "tue", "wed", "thu", "fri", "sat", "sun",
        "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
        "星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日",
    }

    for row_idx in range(1, min(5, sheet.max_row + 1)):
        matches = {}
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip().lower()
                if val in weekday_patterns:
                    matches[col_idx] = val
                # Also match "day1-7" style header
                elif val.startswith("day") and val[3:].isdigit():
                    matches[col_idx] = val

        if len(matches) >= 5:
            cols = sorted(matches.keys())
            return {
                "header_row": row_idx,
                "col_range": (min(cols), max(cols)),
                "num_cols": len(cols),
            }
    return None


class ExcelImporter(BaseImporter):
    """Excel历史数据导入器。"""

    def __init__(self, session=None):
        self._session = session

    @property
    def session(self):
        if self._session is None:
            from daylife.core.database import init_db
            self._session = init_db()
        return self._session

    def probe(self, file_path: str, max_rows: int = 5) -> dict:
        """探测Excel文件结构，使用preview()返回前几条解析结果。"""
        path = Path(file_path)
        year_range = parse_semester_from_filename(path.stem)

        entries = self.preview(file_path)
        samples = []
        for e in entries[:max_rows]:
            samples.append({
                "date": str(e["date"]),
                "content": e["content"][:100],
                "status": e["status"],
                "category": e["category"],
                "sheet": e.get("sheet"),
                "row": e.get("row"),
                "color": e.get("color_info", ""),
            })

        return {
            "file": path.name,
            "year_range": year_range,
            "total_entries": len(entries),
            "samples": samples,
        }

    def preview(self, file_path: str, **kwargs) -> list[dict]:
        """预览导入数据（不写入数据库）。"""
        return self._parse_file(file_path, **kwargs)

    def execute(self, file_path: str, **kwargs) -> ImportResult:
        """执行导入，写入数据库。"""
        from daylife.core.models import Category, DailyEntry, ImportMetadata

        entries = self._parse_file(file_path, **kwargs)
        if not entries:
            return ImportResult(rows_imported=0, rows_skipped=0, errors=["No entries parsed"])

        session = self.session
        rows_imported = 0
        rows_skipped = 0
        errors = []

        # 预加载分类映射
        categories = {c.name: c.id for c in session.query(Category).all()}

        for entry_data in entries:
            entry_date = entry_data["date"]
            content = entry_data["content"]

            # 增量导入：检查是否已存在相同日期+内容的记录
            existing = (
                session.query(DailyEntry)
                .filter(
                    DailyEntry.date == entry_date,
                    DailyEntry.content == content,
                    DailyEntry.source == "excel_import",
                )
                .first()
            )
            if existing:
                rows_skipped += 1
                continue

            category_name = entry_data.get("category", "其他")
            category_id = categories.get(category_name)

            try:
                entry = DailyEntry(
                    date=entry_date,
                    category_id=category_id,
                    content=content,
                    status=entry_data.get("status", "in_progress"),
                    source="excel_import",
                    notes=entry_data.get("notes"),
                    data_json=json.dumps(
                        {
                            "source_file": Path(file_path).name,
                            "source_sheet": entry_data.get("sheet"),
                            "source_row": entry_data.get("row"),
                            "color_info": entry_data.get("color_info"),
                        },
                        ensure_ascii=False,
                    ),
                )
                session.add(entry)
                rows_imported += 1
            except Exception as e:
                errors.append(f"Row {entry_data.get('row', '?')}: {e}")

        # 记录导入元数据
        if rows_imported > 0:
            session.flush()
            dates = [e["date"] for e in entries if e.get("date")]
            meta = ImportMetadata(
                source_file=str(file_path),
                import_type="excel",
                rows_imported=rows_imported,
                rows_skipped=rows_skipped,
                date_range_start=min(dates) if dates else None,
                date_range_end=max(dates) if dates else None,
                notes=f"Imported from {Path(file_path).name}",
            )
            session.add(meta)
            session.commit()

        return ImportResult(
            rows_imported=rows_imported,
            rows_skipped=rows_skipped,
            errors=errors if errors else None,
        )

    def _parse_file(self, file_path: str, **kwargs) -> list[dict]:
        """解析单个Excel文件，返回条目列表。

        支持两种格式：
        1. 日历网格：日期行 + 内容行交替，每行7列（一周）
        2. 星期网格：无日期，表头为 day1-7 或 Mon-Sun，每行为一周
        """
        path = Path(file_path)
        year_range = parse_semester_from_filename(path.stem)
        if not year_range:
            logger.warning("Could not determine year range from filename: %s", path.name)

        wb = openpyxl.load_workbook(str(path), data_only=True)
        all_entries = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row is None or sheet.max_row < 2:
                continue

            # 尝试检测日历网格格式
            calendar = _detect_calendar_grid(sheet)
            if calendar:
                entries = self._parse_calendar_grid(sheet, sheet_name, calendar, year_range)
                all_entries.extend(entries)
                continue

            # 尝试检测星期网格格式（无日期）
            weekday = _detect_weekday_grid(sheet)
            if weekday:
                entries = self._parse_weekday_grid(sheet, sheet_name, weekday, year_range)
                all_entries.extend(entries)
                continue

            logger.info("Skipping sheet '%s': unknown format", sheet_name)

        wb.close()
        return all_entries

    def _parse_calendar_grid(
        self,
        sheet: Worksheet,
        sheet_name: str,
        calendar: dict,
        year_range: tuple[int, int, int, int] | None,
    ) -> list[dict]:
        """解析日历网格格式。

        Pattern:
          Row N:   [date] [date] [date] [date] [date] [date] [date]
          Row N+1: [content] [content] ... (content for each date above)
          Row N+2: (可能有更多内容行，属于同一周)
          Row N+3: [date] [date] ... (下一周)
        """
        date_rows = calendar["date_rows"]
        col_start, col_end = calendar["col_range"]
        entries = []

        # 按行号排序日期行
        sorted_date_rows = sorted(date_rows.keys())

        for i, date_row_idx in enumerate(sorted_date_rows):
            row_dates = date_rows[date_row_idx]

            # 内容行范围：从日期行+1到下一个日期行-1（或sheet末尾）
            content_start = date_row_idx + 1
            if i + 1 < len(sorted_date_rows):
                content_end = sorted_date_rows[i + 1]
            else:
                content_end = sheet.max_row + 1

            # 对每一列（每一天）收集内容
            for col_idx in range(col_start, col_end + 1):
                if col_idx not in row_dates:
                    continue

                raw_date = row_dates[col_idx]
                if year_range:
                    entry_date = correct_date(raw_date, year_range)
                    if not entry_date:
                        continue
                else:
                    entry_date = raw_date

                # 收集该列在内容行中的所有文本
                content_parts = []
                status = "in_progress"
                color_info = None

                for content_row in range(content_start, content_end):
                    cell = sheet.cell(row=content_row, column=col_idx)
                    if cell.value is not None:
                        text = str(cell.value).strip()
                        # 跳过日期值、公式残留、纯数字
                        if isinstance(cell.value, (datetime, date)):
                            continue
                        if text.startswith("=") or text.startswith("<openpyxl"):
                            continue
                        if _is_pure_number(text):
                            continue
                        if _is_skip_text(text):
                            continue
                        if text:
                            content_parts.append(text)
                            # 取第一个有内容的单元格的颜色
                            if len(content_parts) == 1:
                                status = get_cell_status(cell)
                                color_info = describe_color(cell)

                if not content_parts:
                    continue

                content = "、".join(content_parts)

                entries.append({
                    "date": entry_date,
                    "content": content,
                    "status": status,
                    "category": auto_categorize(content),
                    "sheet": sheet_name,
                    "row": content_start,
                    "color_info": color_info,
                })

        return entries

    def _parse_weekday_grid(
        self,
        sheet: Worksheet,
        sheet_name: str,
        weekday: dict,
        year_range: tuple[int, int, int, int] | None,
    ) -> list[dict]:
        """解析星期网格格式（无日期列，每行为一周，列为星期几）。

        需要根据学期范围推算每行对应的日期。
        """
        if not year_range:
            logger.warning("Cannot parse weekday grid without year_range for sheet '%s'", sheet_name)
            return []

        from datetime import timedelta

        header_row = weekday["header_row"]
        col_start, col_end = weekday["col_range"]
        num_cols = weekday["num_cols"]
        entries = []

        # 确定起始日期（学期第一周的周一）
        start_year, start_month = year_range[0], year_range[1]
        start_date = date(start_year, start_month, 1)
        # 调整到该周的周一
        start_date -= timedelta(days=start_date.weekday())

        # 判断列是否包含"day1-7"格式的额外总结列（col8 = "day1-7"之类）
        # 内容列为 col_start 到 col_start + 6 (7天)
        content_cols = list(range(col_start, min(col_start + 7, col_end + 1)))

        # 数据从表头下一行开始
        data_start = header_row + 1
        # 对于"大二上+寒假"格式：每行是一周中各天的同一个活动slot
        # 多行可能属于同一周（同一行=同一个活动，列=不同天）
        # 但更可能的理解是：每行独立，列=不同天

        # 分析：这种格式每行的每列是一天的一个活动
        # 一周可能有多行（不同活动slot）
        # 我们需要把同一天的所有行合并

        # 计算总共有多少数据行
        data_rows = []
        for row_idx in range(data_start, sheet.max_row + 1):
            has_content = False
            for c in content_cols:
                cell = sheet.cell(row=row_idx, column=c)
                if cell.value and str(cell.value).strip():
                    has_content = True
                    break
            if has_content:
                data_rows.append(row_idx)

        if not data_rows:
            return []

        # 估算：按每N行为一周来分组
        # 简单策略：按学期的总周数和总行数来估算每周占几行
        total_weeks = 0
        end_year, end_month = year_range[2], year_range[3]
        end_date = date(end_year, end_month, 28)  # 近似
        total_weeks = max(1, (end_date - start_date).days // 7)

        rows_per_week = max(1, len(data_rows) // total_weeks)
        logger.info(
            "Weekday grid: %d data rows, ~%d weeks, ~%d rows/week",
            len(data_rows), total_weeks, rows_per_week,
        )

        # 按 rows_per_week 分组
        week_idx = 0
        for group_start in range(0, len(data_rows), rows_per_week):
            group_end = min(group_start + rows_per_week, len(data_rows))
            week_start_date = start_date + timedelta(weeks=week_idx)

            for day_offset, col_idx in enumerate(content_cols):
                day_date = week_start_date + timedelta(days=day_offset)

                # 收集这一天在所有行的内容
                day_contents = []
                status = "in_progress"
                color_info = None

                for ri in range(group_start, group_end):
                    row_idx = data_rows[ri]
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value and str(cell.value).strip():
                        day_contents.append(str(cell.value).strip())
                        if len(day_contents) == 1:
                            status = get_cell_status(cell)
                            color_info = describe_color(cell)

                if not day_contents:
                    continue

                content = "、".join(day_contents)
                entries.append({
                    "date": day_date,
                    "content": content,
                    "status": status,
                    "category": auto_categorize(content),
                    "sheet": sheet_name,
                    "row": data_rows[group_start],
                    "color_info": color_info,
                })

            week_idx += 1

        return entries


def import_directory(
    dir_path: str,
    dry_run: bool = False,
    probe_only: bool = False,
    session=None,
) -> dict:
    """批量导入目录下所有xlsx文件。

    Args:
        dir_path: Excel文件所在目录
        dry_run: 仅预览不写入
        probe_only: 仅探测文件结构
        session: 数据库会话

    Returns:
        导入统计结果
    """
    dir_p = Path(dir_path)
    xlsx_files = sorted(dir_p.glob("*.xlsx"))

    if not xlsx_files:
        print(f"No .xlsx files found in {dir_path}")
        return {"files": 0, "total_imported": 0}

    print(f"Found {len(xlsx_files)} Excel files:")
    for f in xlsx_files:
        yr = parse_semester_from_filename(f.stem)
        yr_str = f"{yr[0]}.{yr[1]:02d}-{yr[2]}.{yr[3]:02d}" if yr else "unknown"
        print(f"  - {f.name}  [{yr_str}]")
    print()

    importer = ExcelImporter(session=session)
    total_imported = 0
    total_skipped = 0
    total_errors = []
    results = []

    for xlsx_file in xlsx_files:
        print(f"{'=' * 60}")
        print(f"Processing: {xlsx_file.name}")
        print(f"{'=' * 60}")

        if probe_only:
            probe_result = importer.probe(str(xlsx_file))
            print(f"  Year range: {probe_result['year_range']}")
            print(f"  Total entries: {probe_result['total_entries']}")
            for sample in probe_result["samples"]:
                print(f"    [{sample['date']}] {sample['content'][:60]}  "
                      f"({sample['status']}, {sample['category']})")
            results.append(probe_result)
            continue

        if dry_run:
            entries = importer.preview(str(xlsx_file))
            print(f"  Would import {len(entries)} entries")
            for entry in entries[:5]:
                print(f"    [{entry['date']}] {entry['content'][:50]}  "
                      f"({entry['status']}, {entry['category']})")
            if len(entries) > 5:
                print(f"    ... and {len(entries) - 5} more")
            total_imported += len(entries)
        else:
            result = importer.execute(str(xlsx_file))
            print(f"  Imported: {result.rows_imported}, Skipped: {result.rows_skipped}")
            if result.errors:
                for err in result.errors[:5]:
                    print(f"  Error: {err}")
            total_imported += result.rows_imported
            total_skipped += result.rows_skipped
            if result.errors:
                total_errors.extend(result.errors)

        print()

    print(f"\n{'=' * 60}")
    print(f"{'SUMMARY (DRY RUN)' if dry_run else 'SUMMARY'}")
    print(f"{'=' * 60}")
    print(f"Files processed: {len(xlsx_files)}")
    if dry_run:
        print(f"Total entries to import: {total_imported}")
    else:
        print(f"Total imported: {total_imported}")
        print(f"Total skipped (duplicates): {total_skipped}")
        if total_errors:
            print(f"Total errors: {len(total_errors)}")

    return {
        "files": len(xlsx_files),
        "total_imported": total_imported,
        "total_skipped": total_skipped,
        "total_errors": len(total_errors),
    }


def main():
    parser = argparse.ArgumentParser(
        description="Import Excel historical data into DayLife database"
    )
    parser.add_argument(
        "--path",
        required=True,
        help="Path to Excel file or directory containing .xlsx files",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview import without writing to database",
    )
    parser.add_argument(
        "--probe",
        action="store_true",
        help="Only probe Excel structure (show first few rows)",
    )
    parser.add_argument(
        "--db-path",
        help="Custom database path (default: ~/.local/share/daylife/daylife.db)",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable verbose logging",
    )

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    # 设置数据库路径
    if args.db_path:
        import os
        os.environ["DAYLIFE_DB_PATH"] = args.db_path

    session = None
    if not args.dry_run and not args.probe:
        from daylife.core.database import init_db
        session = init_db()

    target = Path(args.path)
    if target.is_dir():
        import_directory(
            str(target),
            dry_run=args.dry_run,
            probe_only=args.probe,
            session=session,
        )
    elif target.is_file() and target.suffix == ".xlsx":
        importer = ExcelImporter(session=session)
        if args.probe:
            result = importer.probe(str(target))
            print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
        elif args.dry_run:
            entries = importer.preview(str(target))
            print(f"Would import {len(entries)} entries from {target.name}")
            for entry in entries[:10]:
                print(f"  [{entry['date']}] {entry['content'][:60]}  "
                      f"({entry['status']}, {entry['category']})")
            if len(entries) > 10:
                print(f"  ... and {len(entries) - 10} more")
        else:
            result = importer.execute(str(target))
            print(f"Imported: {result.rows_imported}, Skipped: {result.rows_skipped}")
            if result.errors:
                for err in result.errors:
                    print(f"  Error: {err}")
    else:
        print(f"Error: {target} is not a valid file or directory")
        sys.exit(1)


if __name__ == "__main__":
    main()
