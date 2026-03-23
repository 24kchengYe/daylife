"""Check the grid format files more carefully."""
import openpyxl
from daylife.importer.color_parser import get_cell_status, describe_color

# Check 大二上+寒假 for colors
fpath = "D:/my college/zyc学习计划/博一上+寒假.xlsx"
wb = openpyxl.load_workbook(fpath)
sheet = wb['日历']

# Check colors on content rows
print("Color check on 博一上+寒假.xlsx:")
for r in [4, 6, 8, 10]:  # content rows
    for c in range(2, 9):
        cell = sheet.cell(row=r, column=c)
        if cell.value:
            status = get_cell_status(cell)
            color = describe_color(cell)
            print(f"  Row {r}, Col {c}: status={status}, color={color}, content={str(cell.value)[:30]}")
wb.close()

# Check 大二上+寒假 for colors
print("\nColor check on 大二上+寒假.xlsx:")
fpath = "D:/my college/zyc学习计划/大二上+寒假.xlsx"
wb = openpyxl.load_workbook(fpath)
sheet = wb['Sheet1']
for r in range(2, 6):
    for c in range(1, 9):
        cell = sheet.cell(row=r, column=c)
        if cell.value:
            status = get_cell_status(cell)
            color = describe_color(cell)
            print(f"  Row {r}, Col {c}: status={status}, color={color}, content={str(cell.value)[:30]}")
wb.close()

# Check 大四下 for colors
print("\nColor check on 大四下+暑假.xlsx (1月):")
fpath = "D:/my college/zyc学习计划/大四下+暑假.xlsx"
wb = openpyxl.load_workbook(fpath)
sheet = wb['1 月']
for r in [5, 7, 9]:  # content rows
    for c in range(2, 9):
        cell = sheet.cell(row=r, column=c)
        if cell.value:
            status = get_cell_status(cell)
            color = describe_color(cell)
            print(f"  Row {r}, Col {c}: status={status}, color={color}, content={str(cell.value)[:40]}")
wb.close()
