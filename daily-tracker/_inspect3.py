"""Check if colors exist when loading without data_only."""
import openpyxl
from daylife.importer.color_parser import get_cell_status, describe_color

fpath = "D:/my college/zyc学习计划/大三上寒假.xlsx"
wb = openpyxl.load_workbook(fpath, data_only=False)
print(f"Sheets: {wb.sheetnames}")
sheet = wb[wb.sheetnames[0]]
print(f"Sheet: {sheet.title}, rows={sheet.max_row}, cols={sheet.max_column}")

# Check rows for colors
for r in range(1, min(20, sheet.max_row + 1)):
    for c in range(1, min(sheet.max_column + 1, 11)):
        cell = sheet.cell(row=r, column=c)
        if cell.value:
            status = get_cell_status(cell)
            color = describe_color(cell)
            if color != "no-color":
                print(f"  Row {r}, Col {c}: status={status}, color={color}, val={str(cell.value)[:40]}")

# Try deeper rows
print("\nDeeper rows:")
for r in range(20, min(60, sheet.max_row + 1)):
    for c in range(1, min(sheet.max_column + 1, 11)):
        cell = sheet.cell(row=r, column=c)
        if cell.value:
            status = get_cell_status(cell)
            color = describe_color(cell)
            if color != "no-color":
                print(f"  Row {r}, Col {c}: status={status}, color={color}, val={str(cell.value)[:40]}")
wb.close()

# Also check 寒假+博三下 which showed an 'incomplete'
print("\n\nCheck 寒假+博三下+暑假+博四上:")
fpath = "D:/my college/zyc学习计划/寒假+博三下+暑假+博四上（202601-202612）.xlsx"
wb = openpyxl.load_workbook(fpath, data_only=False)
sheet = wb[wb.sheetnames[0]]
for r in range(1, min(30, sheet.max_row + 1)):
    for c in range(1, min(sheet.max_column + 1, 11)):
        cell = sheet.cell(row=r, column=c)
        if cell.value:
            status = get_cell_status(cell)
            color = describe_color(cell)
            if color != "no-color":
                print(f"  Row {r}, Col {c}: status={status}, color={color}, val={str(cell.value)[:40]}")
wb.close()
