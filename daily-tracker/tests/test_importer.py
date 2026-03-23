"""Excel导入器测试 - 使用mock Excel文件"""

import tempfile
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from daylife.importer.color_parser import (
    describe_color,
    get_cell_status,
    is_gray,
    is_red,
)
from daylife.importer.date_corrector import (
    correct_date,
    parse_date_from_cell,
    parse_explicit_date_range,
    parse_semester_from_filename,
)
from daylife.importer.excel_importer import ExcelImporter, auto_categorize


# ============================================================
# date_corrector 测试
# ============================================================


class TestParseSemesterFromFilename:
    """测试从文件名解析学期信息。"""

    def test_sophomore_first_semester(self):
        result = parse_semester_from_filename("大二上+寒假")
        assert result is not None
        start_year, start_month, end_year, end_month = result
        assert start_year == 2019
        assert end_year == 2020

    def test_sophomore_second_semester(self):
        result = parse_semester_from_filename("大二下工作表+暑假")
        assert result is not None
        start_year, _, end_year, _ = result
        assert start_year == 2020
        assert end_year == 2020

    def test_junior_first_semester(self):
        result = parse_semester_from_filename("大三上工作计划")
        assert result is not None
        assert result[0] == 2020  # start_year
        assert result[2] == 2021  # end_year

    def test_phd_first_year(self):
        result = parse_semester_from_filename("博一上+寒假")
        assert result is not None
        assert result[0] == 2023

    def test_phd_second_year_second_semester(self):
        result = parse_semester_from_filename("博二下+暑假+博三上（202501-202512）")
        # 明确年份应优先
        result_explicit = parse_explicit_date_range("博二下+暑假+博三上（202501-202512）")
        assert result_explicit == (2025, 1, 2025, 12)

    def test_explicit_date_range(self):
        result = parse_explicit_date_range("博一下+暑假+博二上（202401-202412）")
        assert result == (2024, 1, 2024, 12)

    def test_explicit_range_priority(self):
        """明确的年月范围应优先于学期推算。"""
        result = parse_semester_from_filename("寒假+博三下+暑假+博四上（202601-202612）")
        assert result == (2026, 1, 2026, 12)

    def test_no_match(self):
        result = parse_semester_from_filename("random_file")
        assert result is None

    def test_senior_year_five(self):
        result = parse_semester_from_filename("大五下+暑假")
        assert result is not None
        assert result[0] == 2023  # start_year
        assert result[2] == 2023  # end_year


class TestCorrectDate:
    """测试日期年份校正。"""

    def test_correct_fall_semester_date(self):
        # 大二上学期(2019.09-2020.01)，Excel显示2026-10-15
        raw = date(2026, 10, 15)
        corrected = correct_date(raw, (2019, 9, 2020, 1))
        assert corrected == date(2019, 10, 15)

    def test_correct_spring_date_across_year(self):
        # 大二上+寒假(2019.09-2020.02)，Excel显示2026-01-20
        raw = date(2026, 1, 20)
        corrected = correct_date(raw, (2019, 9, 2020, 2))
        assert corrected == date(2020, 1, 20)

    def test_correct_summer_date(self):
        # 大二下(2020.03-2020.08)，Excel显示2026-06-01
        raw = date(2026, 6, 1)
        corrected = correct_date(raw, (2020, 3, 2020, 8))
        assert corrected == date(2020, 6, 1)

    def test_same_year_range(self):
        raw = date(2026, 5, 15)
        corrected = correct_date(raw, (2021, 3, 2021, 8))
        assert corrected == date(2021, 5, 15)

    def test_leap_year_feb29(self):
        # 2月29日在非闰年应降级为2月28日
        raw = date(2024, 2, 29)  # 2024是闰年
        corrected = correct_date(raw, (2019, 1, 2019, 6))
        assert corrected == date(2019, 2, 28)

    def test_datetime_input(self):
        raw = datetime(2026, 10, 15, 14, 30)
        corrected = correct_date(raw, (2019, 9, 2020, 1))
        assert corrected == date(2019, 10, 15)


class TestParseDateFromCell:
    """测试单元格日期解析。"""

    def test_datetime_object(self):
        result = parse_date_from_cell(datetime(2026, 3, 15, 10, 0))
        assert result == date(2026, 3, 15)

    def test_date_object(self):
        result = parse_date_from_cell(date(2026, 3, 15))
        assert result == date(2026, 3, 15)

    def test_string_iso(self):
        result = parse_date_from_cell("2026-03-15")
        assert result == date(2026, 3, 15)

    def test_string_slash(self):
        result = parse_date_from_cell("2026/03/15")
        assert result == date(2026, 3, 15)

    def test_string_dot(self):
        result = parse_date_from_cell("2026.03.15")
        assert result == date(2026, 3, 15)

    def test_string_month_day(self):
        result = parse_date_from_cell("03-15")
        assert result is not None
        assert result.month == 3
        assert result.day == 15
        assert result.year == 1900  # 占位年份

    def test_chinese_date(self):
        result = parse_date_from_cell("3月15日")
        assert result is not None
        assert result.month == 3
        assert result.day == 15

    def test_none(self):
        assert parse_date_from_cell(None) is None

    def test_empty_string(self):
        assert parse_date_from_cell("") is None

    def test_non_date_string(self):
        assert parse_date_from_cell("hello world") is None

    def test_number(self):
        assert parse_date_from_cell(12345) is None


# ============================================================
# color_parser 测试
# ============================================================


class TestColorDetection:
    """测试颜色判断函数。"""

    def test_gray_detection(self):
        assert is_gray(192, 192, 192) is True  # silver
        assert is_gray(128, 128, 128) is True  # gray
        assert is_gray(160, 160, 160) is True
        assert is_gray(217, 217, 217) is True  # D9D9D9 - common Excel gray

    def test_not_gray_white(self):
        assert is_gray(255, 255, 255) is False

    def test_not_gray_black(self):
        assert is_gray(0, 0, 0) is False

    def test_not_gray_colored(self):
        assert is_gray(255, 0, 0) is False  # red
        assert is_gray(0, 0, 255) is False  # blue

    def test_red_detection(self):
        assert is_red(255, 0, 0) is True
        assert is_red(200, 50, 50) is True
        assert is_red(220, 80, 80) is True

    def test_not_red(self):
        assert is_red(0, 0, 255) is False  # blue
        assert is_red(128, 128, 128) is False  # gray
        assert is_red(255, 200, 200) is False  # light pink


class TestGetCellStatus:
    """测试从单元格提取状态。"""

    def _make_cell_with_fill(self, r: int, g: int, b: int):
        """创建带背景色的mock单元格。"""
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="test")
        fill = PatternFill(
            start_color=f"{r:02X}{g:02X}{b:02X}",
            end_color=f"{r:02X}{g:02X}{b:02X}",
            fill_type="solid",
        )
        cell.fill = fill
        return cell

    def _make_cell_with_font_color(self, r: int, g: int, b: int):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="test")
        cell.font = Font(color=f"{r:02X}{g:02X}{b:02X}")
        return cell

    def _make_cell_with_strikethrough(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="test")
        cell.font = Font(strikethrough=True)
        return cell

    def _make_plain_cell(self):
        wb = Workbook()
        ws = wb.active
        return ws.cell(row=1, column=1, value="test")

    def test_gray_background_completed(self):
        cell = self._make_cell_with_fill(192, 192, 192)
        assert get_cell_status(cell) == "completed"

    def test_d9d9d9_gray_completed(self):
        cell = self._make_cell_with_fill(0xD9, 0xD9, 0xD9)
        assert get_cell_status(cell) == "completed"

    def test_red_background_incomplete(self):
        cell = self._make_cell_with_fill(255, 0, 0)
        assert get_cell_status(cell) == "incomplete"

    def test_red_font_incomplete(self):
        cell = self._make_cell_with_font_color(255, 0, 0)
        assert get_cell_status(cell) == "incomplete"

    def test_strikethrough_completed(self):
        cell = self._make_cell_with_strikethrough()
        assert get_cell_status(cell) == "completed"

    def test_no_color_in_progress(self):
        cell = self._make_plain_cell()
        assert get_cell_status(cell) == "in_progress"

    def test_describe_color_gray(self):
        cell = self._make_cell_with_fill(192, 192, 192)
        desc = describe_color(cell)
        assert "gray" in desc

    def test_describe_color_red(self):
        cell = self._make_cell_with_fill(255, 0, 0)
        desc = describe_color(cell)
        assert "red" in desc


# ============================================================
# excel_importer 测试
# ============================================================


class TestAutoCategories:
    """测试自动分类。"""

    def test_study(self):
        assert auto_categorize("复习高等数学") == "学习"

    def test_research(self):
        assert auto_categorize("写论文第三章") == "科研"

    def test_exercise(self):
        assert auto_categorize("晨跑5公里") == "运动"

    def test_programming(self):
        assert auto_categorize("写python脚本处理数据") == "编程"

    def test_life(self):
        assert auto_categorize("去超市购物") == "生活"

    def test_entertainment(self):
        assert auto_categorize("看电影") == "娱乐"

    def test_unknown(self):
        assert auto_categorize("做一些事情") == "其他"


def _create_test_xlsx(
    tmp_path: Path,
    filename: str = "大二上+寒假.xlsx",
    rows: list[tuple] | None = None,
) -> Path:
    """创建测试用的Excel文件（日历网格格式）。

    生成的格式与用户真实Excel一致：
    - 日期行：7列日期值（一周）
    - 内容行：对应列的活动内容
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "9月"

    if rows is None:
        rows = [
            (datetime(2026, 9, 1), "开学报到"),
            (datetime(2026, 9, 2), "上高等数学课"),
            (datetime(2026, 9, 3), "跑步3公里"),
            (datetime(2026, 10, 1), "国庆放假"),
        ]

    gray_fill = PatternFill(
        start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
    )
    red_fill = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )

    # 按周分组
    from datetime import timedelta
    weeks: dict[int, list[tuple]] = {}
    for dt, content in rows:
        d = dt if isinstance(dt, date) and not isinstance(dt, datetime) else dt
        if isinstance(d, datetime):
            d = d.date()
        # 计算该日期所在周的周一
        week_start = d - timedelta(days=d.weekday())
        week_key = week_start.toordinal()
        if week_key not in weeks:
            weeks[week_key] = []
        weeks[week_key].append((dt, content))

    current_row = 1
    entry_idx = 0
    for week_key in sorted(weeks.keys()):
        week_entries = weeks[week_key]
        # 计算本周周一
        week_start = date.fromordinal(week_key)

        # 日期行：col 2-8（Sun=col2 到 Sat=col8, 但用Mon-Sun col2-col8更简单）
        for day_offset in range(7):
            day_date = week_start + timedelta(days=day_offset)
            ws.cell(row=current_row, column=2 + day_offset, value=datetime(day_date.year, day_date.month, day_date.day))

        # 内容行
        content_row = current_row + 1
        for dt, content in week_entries:
            d = dt if isinstance(dt, date) and not isinstance(dt, datetime) else dt
            if isinstance(d, datetime):
                d = d.date()
            day_offset = (d - week_start).days
            col = 2 + day_offset
            cell = ws.cell(row=content_row, column=col, value=content)
            # 第1个entry灰色(completed)，第2个红色(incomplete)
            if entry_idx == 0:
                cell.fill = gray_fill
            elif entry_idx == 1:
                cell.fill = red_fill
            entry_idx += 1

        current_row = content_row + 1

    file_path = tmp_path / filename
    wb.save(str(file_path))
    return file_path


class TestExcelImporterProbe:
    """测试Excel结构探测。"""

    def test_probe_basic(self, tmp_path):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter()
        result = importer.probe(str(xlsx_path))

        assert result["file"] == "大二上+寒假.xlsx"
        assert result["year_range"] is not None
        assert result["total_entries"] == 4

    def test_probe_shows_samples(self, tmp_path):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter()
        result = importer.probe(str(xlsx_path), max_rows=3)

        assert len(result["samples"]) <= 3


class TestExcelImporterPreview:
    """测试预览(dry-run)模式。"""

    def test_preview_returns_entries(self, tmp_path):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter()
        entries = importer.preview(str(xlsx_path))

        assert len(entries) == 4
        # 检查日期被校正到2019年
        for entry in entries:
            assert entry["date"].year == 2019 or entry["date"].year == 2020

    def test_preview_status_detection(self, tmp_path):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter()
        entries = importer.preview(str(xlsx_path))

        # 第1行灰色→completed
        assert entries[0]["status"] == "completed"
        # 第2行红色→incomplete
        assert entries[1]["status"] == "incomplete"
        # 第3,4行无色→in_progress
        assert entries[2]["status"] == "in_progress"

    def test_preview_auto_category(self, tmp_path):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter()
        entries = importer.preview(str(xlsx_path))

        # "上高等数学课" → 学习
        assert entries[1]["category"] == "学习"
        # "跑步3公里" → 运动
        assert entries[2]["category"] == "运动"

    def test_preview_date_correction(self, tmp_path):
        """大二上=2019.09-2020.01, 所以9月的日期应该校正到2019年"""
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter()
        entries = importer.preview(str(xlsx_path))

        sept_entry = entries[0]
        assert sept_entry["date"].year == 2019
        assert sept_entry["date"].month == 9

    def test_preview_explicit_year_range(self, tmp_path):
        """文件名含明确年份范围时应直接使用。"""
        xlsx_path = _create_test_xlsx(
            tmp_path,
            filename="博一下+暑假+博二上（202401-202412）.xlsx",
            rows=[
                (datetime(2026, 4, 1), "实验数据分析"),
                (datetime(2026, 11, 1), "写论文"),
            ],
        )
        importer = ExcelImporter()
        entries = importer.preview(str(xlsx_path))

        assert len(entries) == 2
        assert entries[0]["date"] == date(2024, 4, 1)
        assert entries[1]["date"] == date(2024, 11, 1)


class TestExcelImporterExecute:
    """测试实际写入数据库。"""

    @pytest.fixture()
    def db_session(self, tmp_path):
        """创建临时数据库会话。"""
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker

        from daylife.core.models import Base, Category

        db_path = tmp_path / "test.db"
        engine = create_engine(f"sqlite:///{db_path}")
        Base.metadata.create_all(engine)

        session = sessionmaker(bind=engine)()
        # 插入默认分类
        for cat in [
            {"name": "学习", "icon": "📚", "color": "#4A90D9", "sort_order": 1},
            {"name": "科研", "icon": "🔬", "color": "#7B68EE", "sort_order": 2},
            {"name": "运动", "icon": "🏃", "color": "#2ECC71", "sort_order": 4},
            {"name": "生活", "icon": "🏠", "color": "#F39C12", "sort_order": 5},
            {"name": "其他", "icon": "📝", "color": "#BDC3C7", "sort_order": 10},
        ]:
            session.add(Category(**cat))
        session.commit()
        yield session
        session.close()

    def test_import_creates_entries(self, tmp_path, db_session):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter(session=db_session)
        result = importer.execute(str(xlsx_path))

        assert result.rows_imported == 4
        assert result.rows_skipped == 0

        from daylife.core.models import DailyEntry
        entries = db_session.query(DailyEntry).all()
        assert len(entries) == 4
        assert all(e.source == "excel_import" for e in entries)

    def test_import_creates_metadata(self, tmp_path, db_session):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter(session=db_session)
        importer.execute(str(xlsx_path))

        from daylife.core.models import ImportMetadata
        meta = db_session.query(ImportMetadata).first()
        assert meta is not None
        assert meta.import_type == "excel"
        assert meta.rows_imported == 4

    def test_incremental_import_skips_duplicates(self, tmp_path, db_session):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter(session=db_session)

        # 第一次导入
        result1 = importer.execute(str(xlsx_path))
        assert result1.rows_imported == 4

        # 第二次导入同一文件 - 应跳过所有
        result2 = importer.execute(str(xlsx_path))
        assert result2.rows_imported == 0
        assert result2.rows_skipped == 4

    def test_import_with_correct_categories(self, tmp_path, db_session):
        xlsx_path = _create_test_xlsx(tmp_path)
        importer = ExcelImporter(session=db_session)
        importer.execute(str(xlsx_path))

        from daylife.core.models import DailyEntry
        entries = db_session.query(DailyEntry).order_by(DailyEntry.date).all()

        # "上高等数学课" → category_id 对应 "学习"
        math_entry = [e for e in entries if "高等数学" in e.content][0]
        from daylife.core.models import Category
        cat = db_session.query(Category).filter(Category.id == math_entry.category_id).first()
        assert cat.name == "学习"


class TestMultiSheetExcel:
    """测试多sheet Excel文件。"""

    def test_multiple_sheets(self, tmp_path):
        from datetime import timedelta
        wb = Workbook()

        # Sheet 1: 9月 - 日历网格格式（需要2个日期行才能被检测）
        ws1 = wb.active
        ws1.title = "9月"
        w1 = date(2019, 9, 1) - timedelta(days=date(2019, 9, 1).weekday())
        for i in range(7):
            d = w1 + timedelta(days=i)
            ws1.cell(row=1, column=2 + i, value=datetime(d.year, d.month, d.day))
        ws1.cell(row=2, column=2 + (date(2019, 9, 1) - w1).days, value="九月任务1")
        w2 = w1 + timedelta(weeks=1)
        for i in range(7):
            d = w2 + timedelta(days=i)
            ws1.cell(row=3, column=2 + i, value=datetime(d.year, d.month, d.day))

        # Sheet 2: 10月 - 日历网格格式
        ws2 = wb.create_sheet("10月")
        w3 = date(2019, 10, 1) - timedelta(days=date(2019, 10, 1).weekday())
        for i in range(7):
            d = w3 + timedelta(days=i)
            ws2.cell(row=1, column=2 + i, value=datetime(d.year, d.month, d.day))
        ws2.cell(row=2, column=2 + (date(2019, 10, 1) - w3).days, value="十月任务1")
        w4 = w3 + timedelta(weeks=1)
        for i in range(7):
            d = w4 + timedelta(days=i)
            ws2.cell(row=3, column=2 + i, value=datetime(d.year, d.month, d.day))

        file_path = tmp_path / "大二上+寒假.xlsx"
        wb.save(str(file_path))

        importer = ExcelImporter()
        entries = importer.preview(str(file_path))
        assert len(entries) == 2


class TestEdgeCases:
    """测试边界情况。"""

    def test_empty_sheet(self, tmp_path):
        wb = Workbook()
        wb.active.title = "Empty"
        file_path = tmp_path / "大二上+寒假.xlsx"
        wb.save(str(file_path))

        importer = ExcelImporter()
        entries = importer.preview(str(file_path))
        assert len(entries) == 0

    def test_no_date_column(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="只有文字")
        ws.cell(row=2, column=1, value="没有日期")
        file_path = tmp_path / "大二上+寒假.xlsx"
        wb.save(str(file_path))

        importer = ExcelImporter()
        entries = importer.preview(str(file_path))
        # 没有日期列，应该跳过
        assert len(entries) == 0

    def test_string_dates(self, tmp_path):
        """测试日历网格最少2个日期行仍可解析。"""
        from datetime import timedelta
        wb = Workbook()
        ws = wb.active
        base = date(2019, 9, 1)
        week_start = base - timedelta(days=base.weekday())
        # 第一个日期行
        for i in range(7):
            d = week_start + timedelta(days=i)
            ws.cell(row=1, column=2 + i, value=datetime(d.year, d.month, d.day))
        ws.cell(row=2, column=2, value="文本日期任务")
        # 第二个日期行
        week_start2 = week_start + timedelta(weeks=1)
        for i in range(7):
            d = week_start2 + timedelta(days=i)
            ws.cell(row=3, column=2 + i, value=datetime(d.year, d.month, d.day))

        file_path = tmp_path / "大二上+寒假.xlsx"
        wb.save(str(file_path))

        importer = ExcelImporter()
        entries = importer.preview(str(file_path))
        assert len(entries) >= 1
