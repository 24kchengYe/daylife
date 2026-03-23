"""Check why 大三上工作计划 fails detection."""
import openpyxl
from datetime import datetime, date

for fpath in [
    "D:/my college/zyc学习计划/大三上工作计划.xlsx",
    "D:/my college/zyc学习计划/大四上工作计划+寒假.xlsx",
    "D:/my college/zyc学习计划/大三下工作计划+暑假.xlsx",
]:
    print(f"\n{'='*60}")
    print(f"FILE: {fpath}")
    wb = openpyxl.load_workbook(fpath, data_only=False)
    sheet = wb[wb.sheetnames[0]]
    print(f"Sheet: {sheet.title}, rows={sheet.max_row}, cols={sheet.max_column}")

    # Check first 25 rows for dates
    for r in range(1, min(25, sheet.max_row + 1)):
        date_count = 0
        parts = []
        for c in range(1, min(sheet.max_column + 1, 11)):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                if isinstance(val, (datetime, date)):
                    parts.append(f"c{c}=[DATE:{val}]")
                    date_count += 1
                else:
                    s = str(val)[:30]
                    parts.append(f"c{c}={s}")
        if parts:
            marker = " *** DATE ROW" if date_count >= 3 else ""
            print(f"  Row {r:2d}: {' | '.join(parts)}{marker}")
    wb.close()
